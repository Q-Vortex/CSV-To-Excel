#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import argparse
import re

# --- ПАРСИНГ АРГУМЕНТОВ ---
parser = argparse.ArgumentParser(description="Запись CSV данных в Excel файл.")
parser.add_argument("csv_name", help="Имя CSV файла для ввода (с расширением)")
parser.add_argument("sep", help="Разделитель в CSV файле, например , или ;")
parser.add_argument("excel_name", help="Имя Excel файла для вывода (с расширением)")
parser.add_argument("--sheet_name", help="Имя листа в Excel для вывода (если не указано, генерируется)", default="")
args = parser.parse_args()

csv_name = args.csv_name
sep = args.sep
excel_name = args.excel_name
sheet_name = args.sheet_name.strip()

base_name = "Sheet"

# --- ОТКРЫТИЕ ФАЙЛОВ ---
try:
    wb = openpyxl.load_workbook(excel_name)
except FileNotFoundError:
    print(f"Файл '{excel_name}' не найден. Создаю новый.")
    wb = openpyxl.Workbook()

# --- ОПРЕДЕЛЕНИЕ ИМЕНИ ЛИСТА ---
if sheet_name == "":
    pattern = re.compile(rf"^{re.escape(base_name)}(\d*)$")
    max_num = 0
    for name in wb.sheetnames:
        m = pattern.match(name)
        if m:
            num_str = m.group(1)
            if num_str == "":
                num = 0
            else:
                num = int(num_str)
            if num > max_num:
                max_num = num
    if max_num == 0 and base_name not in wb.sheetnames:
        new_sheet_name = base_name
    else:
        new_sheet_name = f"{base_name}{max_num + 1}"
    sheet_name = new_sheet_name

if sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
else:
    print(f"Лист '{sheet_name}' не найден. Создаю новый лист.")
    sheet = wb.create_sheet(sheet_name)

# --- ОТКРЫТИЕ CSV ФАЙЛА ---
try:
    file = open(csv_name, "r", encoding="utf-8")
except Exception as e:
    print(f"Ошибка при открытии CSV файла: {e}")
    sys.exit(1)

# --- ЗАПИСЬ ДАННЫХ ---
row = 1
for line in file:
    line = line.rstrip("\n").split(sep)
    column = 1
    for data in line:
        sheet.cell(row, column).value = data
        column += 1
    row += 1

# --- СОХРАНЕНИЕ ---
wb.save(excel_name)
file.close()

print(f"Данные из '{csv_name}' записаны в файл '{excel_name}', лист '{sheet_name}'")
